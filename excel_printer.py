"""
Known issues:
    - Dataframes with duplicate colnames can produce unexpected results without crashing
"""

from pandas import DataFrame
import openpyxl
from openpyxl.utils import get_column_letter, get_column_interval
from openpyxl.styles import Alignment, Font, PatternFill
from os.path import exists
import warnings
warnings.simplefilter(action='ignore', category=UserWarning)


def expand_range(val):
    if ':' not in val:
        return [val]
    l = val.split(':')
    return get_column_interval(l[0], l[1])

def listify(val):
    '''Returns val in list if not already a list'''
    if not isinstance(val, list):
        return [val]
    else:
        return val

class Printer():
    '''
    Make instance and add frames to be printed with function "append"
    parameters:
        auto_fmt_col_width: def. False
        wrap_cols: def. False
        float_as_percent. def. False
        col_width_dict
    '''

    SCALE_FACTOR = .6      # For finding col-width
    percent_format = '#0%'
    # float_format = '#0.0'

    def __init__(self, path='output.xlsx', overwrite_sheet=True, **kwargs):
        kwargs.update(kwargs.get('kwargs', {}))
        if not path.endswith('.xlsx'):
            path = path + '.xlsx'
        if exists(path):
            self.wb = openpyxl.load_workbook(path)     # Opens existing sheet
        else:
            self.wb = openpyxl.Workbook()                   # Creates new sheet
        self.path = path
        self.que = {}
        self.col_width_dict = {}
        self.custom_col_width = {}
        self.overwrite_sheet = overwrite_sheet
        self.auto_fmt_col_width = kwargs.get('auto_fmt_col_width', False)
        self.wrap_cols = kwargs.get('wrap_cols', False)
        self.float_as_percent = kwargs.get('float_as_percent', False)
        self.index_dict = {}

    def set_sheet(self, target_sheet):
        if not target_sheet in self.wb.sheetnames:
            self.wb.create_sheet(target_sheet)
        else:
            if self.overwrite_sheet:
                del self.wb[target_sheet]
                self.wb.create_sheet(target_sheet)
        self.sh = self.wb[target_sheet]

    def append(self, df, sheetname='Sheet', custom_col_width={}, index=True, header=True, percent_cols=[], float_fmt='', wrap_values=False, **kwargs):
        '''Adds dataframe to que.  
        parameters:
            sheetname: str optional
            custom_col_width: {col(int/str):width}
            index: bool
            percent_cols: list or str "skip" to force no float to percent in excel
            wrap_values: bool
            color_dict: {val: 'color'}
            hyperlink_cols: []
        '''
        kwargs.update(kwargs.get('kwargs', {}))
        percent_cols = kwargs.get('percent_cols', percent_cols)
        custom_col_width = kwargs.get('custom_col_width', custom_col_width)
        wrap_values = kwargs.get('wrap_values', wrap_values)
        

        if not isinstance(df, DataFrame):
            return

        self.index_dict.update(())
        for k, v in custom_col_width.items():                   # Custom col width
            self.custom_col_width.update({(k, sheetname): v})

        # Adding attributes to df
        df.incl_idx = index
        df.incl_header = header
        df.percent_cols = percent_cols
        df.float_fmt = float_fmt
        df.wrap_values = wrap_values
        df.color_dict = kwargs.get('color_dict', {})
        df.hyperlink_cols = kwargs.get('hyperlink_cols', [])


        if sheetname in self.que:
            self.que[sheetname].append(df)
            return
        self.que.update({sheetname: [df]})

    def write_to_file(self, df, sheetname, startrow=None, startcol=None):
        index_add = int(df.incl_idx)-1
        sh = self.sh
        if not startrow:
            if sh.max_row > 1:
                startrow = sh.max_row + 1
            else:
                startrow = 0
        if df.incl_idx:
            colstart = len(df.index.names)
        else:
            colstart = 1
            last_col = 1
        has_index_extra = 0
        percent_cols = []
        [percent_cols.extend(expand_range(c)) for c in listify(df.percent_cols)]
        # print(percent_cols)

        if df.incl_header:
            # Printing names of columns (might be multiindex columns)
            for i, colname in enumerate(df.columns.names):
                sh.cell(i+1 + startrow, colstart).value = (colname)
                self.format_column(sh.cell(i+1 + startrow, colstart))
                last_row = i+1
        else:
            last_row = 0

        if df.incl_idx:
            # Printing indexnames
            for col, indexname in enumerate(df.index.names):
                if col or indexname:
                    has_index_extra = 1
                cell = sh.cell(last_row+1 + startrow, col+1)
                cell.value = indexname
                self.format_column(cell)
                last_col = col+1
            # printing index
            for i, element in enumerate(df.index):
                if len(df.index.names) > 1:
                    for ii, sub_element in enumerate(element):
                        if i > 0:
                            if sub_element in df.index[i-1]:
                                # print(f'skipping {sub_element}')
                                continue
                        col = ii + 1
                        row = i+1 + last_row+1
                        cell = sh.cell(row + startrow, col)
                        cell.value = (sub_element)
                else:
                    cell = sh.cell(last_row + i+1 + has_index_extra +startrow, 1)
                    cell.value = (element)
                self.format_index(cell)            # vanlig markering index 
                
        if df.incl_header:
            # Printing columns
            for i, column in enumerate(df.columns):
                col_idx = last_col + i+1 + index_add
                if len(df.columns.names) > 1:                           # Multi column
                    for ii, element in enumerate(column):
                        if i > 0:
                            # Dont duplicate values
                            if element in df.columns[i-1]:
                                continue
                        self.check_col_len(
                            col_string=element, col_idx=col_idx, sheetname=sheetname)
                        cell = sh.cell(ii+1 + startrow, col_idx)
                        cell.value = (element)
                else:
                    cell = sh.cell(1 + startrow, col_idx)
                    cell.value = (column)  # TODO: Make same for index
                    self.check_col_len(col_string=column,
                                       col_idx=col_idx, sheetname=sheetname)
                self.format_column(cell)

                                                                        # printing values
        for i, col in enumerate(df):
            for ii, val in enumerate(df[col]):
                col = last_col + i+1 + index_add
                col_let = get_column_letter(col)
                row = last_row + ii+1 + has_index_extra
                cell = sh.cell(row + startrow, col)
                
                if val in df.color_dict:    # Fulhackar in färg
                    cell.fill = PatternFill(
                        start_color=df.color_dict[val]['color'],
                        end_color=df.color_dict[val]['color'],
                        fill_type="solid"
                        )
                elif col_let in df.hyperlink_cols:
                    cell.hyperlink = (val)
                    cell.value= ('Länk')
                else:
                    cell.value = (val)

                if (self.float_as_percent or percent_cols) and not 'skip' in [str(e).lower() for e in percent_cols]:
                    if not percent_cols:
                        pass
                    elif percent_cols and col_let not in percent_cols:
                        continue
                    
                    cell.number_format = self.percent_format
                
                if df.float_fmt:
                    cell.number_format = df.float_fmt
                if df.wrap_values:
                    cell.alignment = Alignment(wrap_text=True)


    def check_col_len(self, col_string, col_idx, sheetname):
        if not self.auto_fmt_col_width:
            return
        len_col = len(str(col_string))
        if len_col < 13:
            return
        width = int(len(col_string) * self.SCALE_FACTOR)
        # Only update if value is greater than prev val
        if (col_idx, sheetname) in self.col_width_dict:
            if width < self.col_width_dict.get((col_idx, sheetname)):
                return
        print(f'setting col {col_idx} to {width}')
        self.col_width_dict.update({(col_idx, sheetname): width})

    def format_column(self, cell):
        if self.wrap_cols:
            cell.alignment = Alignment(wrap_text=True)
        cell.font = Font(bold=True)

    def format_index(self, cell):
        cell.font = Font(bold=False)

    def format_column_width(self):
        if self.auto_fmt_col_width:
            for col_idx_sheet, width in self.col_width_dict.items():  # col_idx_sheet = (col_idx, sheetname)
                sh = self.wb[col_idx_sheet[1]]
                sh.column_dimensions[get_column_letter(
                    col_idx_sheet[0])].width = width
        for col_idx_sheet, width in self.custom_col_width.items():
            sh = self.wb[col_idx_sheet[1]]
            if isinstance(col_idx_sheet[0], int):
                let = get_column_letter(col_idx_sheet[0])
                sh.column_dimensions[let].width = width
                continue
            let = col_idx_sheet[0]
            if ':' in let:
                for l in expand_range(let):
                    sh.column_dimensions[l].width = width
            else:
                sh.column_dimensions[let].width = width

    def save(self):
        # Check if we should delete default sheet
        if not 'Sheet' in self.que and 'Sheet' in self.wb.sheetnames and len(self.wb.sheetnames) > 1:
            del self.wb['Sheet']
        self.wb.save(self.path)


    def run(self):
        for sheetname, df_list in self.que.items():                 # Run though qued df's and write to workbook
            self.set_sheet(sheetname)
            for df in df_list:
                self.write_to_file(df, sheetname)
        self.format_column_width()
        self.save()

if __name__ == '__main__':
    pass
